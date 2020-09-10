import openpyxl


def getRowCount(file,sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    maxrow = sheet.max_row
    return maxrow

def retColumnCount(file,sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    maxcolumn = sheet.max_column
    return maxcolumn

def readData(file,sheetname,rownum,colnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    readexcel = sheet.cell(row=rownum,column=colnum).value
    return readexcel

def writeData(file,sheetname,rownum,colnum,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    sheet.cell(row=rownum,column=colnum,value=data)
    workbook.save(file)